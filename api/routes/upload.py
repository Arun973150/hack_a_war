"""
File upload endpoint for control onboarding.
CSV/XLSX: direct column parsing.
PDF/DOCX: text extraction → Gemini extracts controls as JSON.
"""
import csv
import io
import json
import structlog
import vertexai
from fastapi import APIRouter, UploadFile, File, HTTPException
from langchain_google_vertexai import ChatVertexAI
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import JsonOutputParser

from config import settings
from knowledge.graph.neo4j_client import Neo4jClient

logger = structlog.get_logger()
router = APIRouter()

vertexai.init(project=settings.vertex_project, location=settings.vertex_location)

EXTRACT_CONTROLS_PROMPT = ChatPromptTemplate.from_messages([
    ("system", """You are a compliance expert. Extract compliance controls from the document text.
Return JSON: {{"controls": [{{"control_id": "...", "name": "...", "description": "...", "framework": "...", "owner_team": "..."}}]}}
If no control_id is obvious, generate one like CTRL-001. Framework should be one of: ISO27001, SOC2, GDPR, PCI-DSS, HIPAA, DORA, NIST, or the most relevant one."""),
    ("human", "Extract compliance controls from this document:\n\n{text}"),
])


def _parse_csv(content: bytes) -> list[dict]:
    reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace")))
    controls = []
    for row in reader:
        # Normalize common column names
        control = {
            "control_id": row.get("control_id") or row.get("id") or row.get("ID") or f"CTRL-{len(controls)+1:03d}",
            "name": row.get("name") or row.get("control_name") or row.get("Name") or "Unnamed Control",
            "description": row.get("description") or row.get("Description") or "",
            "framework": row.get("framework") or row.get("Framework") or "UNKNOWN",
            "owner_team": row.get("owner") or row.get("owner_team") or row.get("Owner") or "Compliance",
            "coverage_score": float(row.get("coverage_score") or row.get("coverage") or 0.0),
        }
        controls.append(control)
    return controls


def _parse_xlsx(content: bytes) -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).lower().strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        controls = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_dict = dict(zip(headers, row))
            controls.append({
                "control_id": str(row_dict.get("control_id") or row_dict.get("id") or f"CTRL-{len(controls)+1:03d}"),
                "name": str(row_dict.get("name") or row_dict.get("control_name") or "Unnamed"),
                "description": str(row_dict.get("description") or ""),
                "framework": str(row_dict.get("framework") or "UNKNOWN"),
                "owner_team": str(row_dict.get("owner") or row_dict.get("owner_team") or "Compliance"),
                "coverage_score": float(row_dict.get("coverage_score") or 0.0),
            })
        return controls
    except Exception as e:
        raise HTTPException(400, f"XLSX parse error: {e}")


def _extract_text_from_pdf(content: bytes) -> str:
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(content))
        return "\n".join(page.extract_text() or "" for page in reader.pages[:20])
    except Exception as e:
        raise HTTPException(400, f"PDF extraction error: {e}")


def _extract_text_from_docx(content: bytes) -> str:
    try:
        import docx
        doc = docx.Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception as e:
        raise HTTPException(400, f"DOCX extraction error: {e}")


def _ai_extract_controls(text: str) -> list[dict]:
    llm = ChatVertexAI(
        model_name=settings.gemini_flash_model,
        project=settings.vertex_project,
        location=settings.vertex_location,
        temperature=0.0,
        max_output_tokens=4096,
    )
    chain = EXTRACT_CONTROLS_PROMPT | llm | JsonOutputParser()
    result = chain.invoke({"text": text[:8000]})
    return result.get("controls", [])


_STATUS_TO_COVERAGE = {
    "implemented": 1.0,
    "operational": 1.0,
    "approved": 0.8,
    "partial": 0.5,
    "draft": 0.3,
    "not implemented": 0.0,
    "non-compliant": 0.0,
}

_FRAMEWORK_ALIASES = {
    "pci_dss": "PCI_DSS", "pci-dss": "PCI_DSS", "pcidss": "PCI_DSS",
    "gdpr": "GDPR",
    "iso_27001": "ISO27001", "iso27001": "ISO27001", "iso 27001": "ISO27001",
    "soc_2": "SOC2", "soc2": "SOC2",
    "hipaa": "HIPAA",
    "eu_ai_act": "EU_AI_ACT", "eu ai act": "EU_AI_ACT",
    "dora": "DORA",
    "sebi": "SEBI",
    "rbi": "RBI", "rbi_payment_guidelines": "RBI",
    "dpdp_india": "GDPR",  # closest match in current enum
}


def _normalize_framework(raw: str) -> str:
    return _FRAMEWORK_ALIASES.get(raw.lower().strip(), "CUSTOM")


def _register_controls(controls: list[dict]) -> list[dict]:
    neo4j = Neo4jClient()
    registered = []
    for c in controls:
        try:
            neo4j.upsert_control({
                "id": c.get("control_id", ""),
                "name": c.get("name", ""),
                "description": c.get("description", ""),
                "owner_team": c.get("owner_team", "Compliance"),
                "framework": c.get("framework", "UNKNOWN"),
                "coverage_score": float(c.get("coverage_score", 0.0)),
            })
            registered.append({**c, "status": "registered"})
        except Exception as e:
            registered.append({**c, "status": "error", "error": str(e)})
    return registered


def _finish_import(controls: list[dict], method: str, filename: str) -> dict:
    if not controls:
        return {"registered": [], "total": 0, "method": method, "message": "No controls found in file"}
    registered = _register_controls(controls)
    logger.info("controls_imported", filename=filename, total=len(registered), method=method)
    return {"registered": registered, "total": len(registered), "method": method, "filename": filename}


def _import_org_profile_json(content: bytes, filename: str) -> dict:
    """
    Parse a Red Forge org profile JSON file (like test.json).
    Handles:
      - controls[]  → registered into Neo4j + PostgreSQL
      - regulations[]  → seeded into regulation_tracking
      - compliance_status{}  → stored as regulation coverage
      - knowledge_graph{nodes, edges}  → pushed to Neo4j
      - company / infrastructure  → stored as org context metadata
    """
    try:
        data = json.loads(content.decode("utf-8", errors="replace"))
    except Exception as e:
        raise HTTPException(400, f"Invalid JSON: {e}")

    results = {
        "filename": filename,
        "method": "org_profile_json",
        "company": data.get("company", {}).get("name", "Unknown"),
        "controls_registered": [],
        "regulations_seeded": [],
        "graph_nodes": 0,
        "graph_edges": 0,
        "compliance_status": {},
    }

    # ── 1. Import controls ──────────────────────────────────────────────────
    raw_controls = data.get("controls", [])
    if raw_controls:
        normalized = []
        for c in raw_controls:
            status_str = str(c.get("status", "")).lower()
            coverage = _STATUS_TO_COVERAGE.get(status_str, 0.5)
            frameworks = c.get("mapped_to", [])
            framework = _normalize_framework(frameworks[0]) if frameworks else "CUSTOM"

            normalized.append({
                "control_id": c.get("id", ""),
                "name": c.get("name", ""),
                "description": f"Mapped to: {', '.join(frameworks)}. Version: {c.get('version', '')}",
                "framework": framework,
                "owner_team": "Compliance",
                "coverage_score": coverage,
            })

        results["controls_registered"] = _register_controls(normalized)

    # ── 2. Seed regulation tracking ─────────────────────────────────────────
    from org_context.models.database import upsert_regulation_tracking
    compliance_status = data.get("compliance_status", {})
    regulations = data.get("regulations", [])
    company_name = data.get("company", {}).get("name", "unknown")

    for reg_name in regulations:
        status = compliance_status.get(reg_name, "Unknown")
        is_relevant = status in ("Partial", "Non-Compliant")
        risk = 8 if status == "Non-Compliant" else (5 if status == "Partial" else 2)
        try:
            upsert_regulation_tracking(
                source_id=f"org-profile-{company_name}-{reg_name}",
                title=reg_name.replace("_", " "),
                jurisdiction=_infer_jurisdiction(reg_name),
                regulatory_body=reg_name,
                document_type="regulation",
                published_date="2024-01-01",
                source_url="",
                is_relevant=is_relevant,
                relevance_score=0.9 if is_relevant else 0.3,
                overall_risk_score=risk,
                impact_summary=f"Compliance status: {status}",
                processing_status="seeded",
            )
            results["regulations_seeded"].append({"name": reg_name, "status": status})
        except Exception as e:
            results["regulations_seeded"].append({"name": reg_name, "status": "error", "error": str(e)})

    results["compliance_status"] = compliance_status

    # ── 3. Push knowledge graph to Neo4j ────────────────────────────────────
    kg = data.get("knowledge_graph", {})
    nodes = kg.get("nodes", [])
    edges = kg.get("edges", [])

    if nodes or edges:
        try:
            neo4j = Neo4jClient()
            for node in nodes:
                try:
                    neo4j.upsert_control({
                        "id": node, "name": node,
                        "description": f"Org profile node: {node}",
                        "owner_team": "System", "framework": "CUSTOM",
                        "coverage_score": 0.0,
                    })
                except Exception:
                    pass
            results["graph_nodes"] = len(nodes)

            # Store edges as relationships
            for edge in edges:
                if len(edge) == 3:
                    try:
                        neo4j.create_relationship(edge[0], edge[2], edge[1])
                    except Exception:
                        pass
            results["graph_edges"] = len(edges)
        except Exception as e:
            logger.warning("neo4j_org_profile_import_failed", error=str(e))

    logger.info(
        "org_profile_imported",
        filename=filename,
        controls=len(results["controls_registered"]),
        regulations=len(results["regulations_seeded"]),
    )
    return results


def _infer_jurisdiction(reg_name: str) -> str:
    reg_lower = reg_name.lower()
    if any(x in reg_lower for x in ("gdpr", "eu_ai", "dora", "nis")):
        return "EU"
    if any(x in reg_lower for x in ("rbi", "dpdp", "sebi", "india")):
        return "India"
    if any(x in reg_lower for x in ("sec", "hipaa", "soc", "ccpa")):
        return "US"
    return "Global"


@router.post("/import-controls")
async def import_controls(file: UploadFile = File(...)):
    """
    Upload a control registry file. Supported: CSV, XLSX (direct parse), PDF, DOCX (AI extraction).
    Registers all controls in Neo4j knowledge graph.
    """
    content = await file.read()
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext == "csv":
        controls = _parse_csv(content)
        method = "direct_parse"
        return _finish_import(controls, method, filename)
    elif ext in ("xlsx", "xls"):
        controls = _parse_xlsx(content)
        method = "direct_parse"
        return _finish_import(controls, method, filename)
    elif ext == "pdf":
        text = _extract_text_from_pdf(content)
        controls = _ai_extract_controls(text)
        method = "ai_extraction"
        return _finish_import(controls, method, filename)
    elif ext in ("docx", "doc"):
        text = _extract_text_from_docx(content)
        controls = _ai_extract_controls(text)
        method = "ai_extraction"
        return _finish_import(controls, method, filename)
    elif ext == "json":
        return _import_org_profile_json(content, filename)
    else:
        raise HTTPException(400, f"Unsupported file type: .{ext}. Supported: csv, xlsx, pdf, docx, json")
